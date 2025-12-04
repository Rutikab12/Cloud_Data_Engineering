"""
Airflow DAG: create_summary_sheet
Reads Sheet1 from an Excel workbook, creates a grouped SUM summary (pivot-like),
appends a Grand Total row, and writes a Summary sheet back to the same workbook.

Requirements (install on Airflow worker):
  pip install pandas openpyxl

Notes:
 - Make sure the worker has access to the file path (EXCEL_PATH).
 - To avoid corruption, the DAG writes to a temporary file then atomically replaces the original.
"""

from datetime import datetime
import logging
import pandas as pd
import tempfile
import shutil
from pathlib import Path
from openpyxl import load_workbook

from airflow import DAG
from airflow.operators.python import PythonOperator

# ---- Config ----
EXCEL_PATH_STR = r"C:\Users\hp\Downloads\Sample_data.xlsx"  # update if needed on worker
SHEET_NAME = "Sheet1"
SUMMARY_SHEET = "Summary"

DIMENSIONS = ["Channel_name", "Group_Dealer_Code", "Group_Dealer_Name", "Commission_Scheme"]
MEASURES = ["Total_Commission_Bonus", "IMEI"]  # aggregate with SUM

default_args = {
    "owner": "airflow",
    "depends_on_past": False,
    "retries": 1,
}

def create_summary_sheet(excel_path: str, sheet_name: str = SHEET_NAME, summary_sheet: str = SUMMARY_SHEET):
    log = logging.getLogger("airflow.task")
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")

    # 1) Read the source sheet
    log.info("Reading Excel file: %s (sheet: %s)", excel_path, sheet_name)
    df = pd.read_excel(excel_path, sheet_name=sheet_name)

    # 2) Validate columns
    missing_dims = [c for c in DIMENSIONS if c not in df.columns]
    missing_meas = [c for c in MEASURES if c not in df.columns]
    if missing_dims or missing_meas:
        raise ValueError(f"Missing columns. Missing dimensions: {missing_dims}, missing measures: {missing_meas}")

    # 3) Ensure measures numeric; coerce non-numeric -> 0
    for m in MEASURES:
        df[m] = pd.to_numeric(df[m], errors="coerce").fillna(0)

    # 4) Group by (pivot-like) and SUM
    log.info("Grouping by dimensions and summing measures...")
    summary = df.groupby(DIMENSIONS, dropna=False)[MEASURES].sum().reset_index()

    # 5) Compute grand totals and append as last row
    grand_totals = {dim: "" for dim in DIMENSIONS}
    grand_totals[DIMENSIONS[0]] = "Grand Total"
    for m in MEASURES:
        # ensure numeric sum
        grand_totals[m] = float(summary[m].sum()) if not summary.empty else 0.0

    summary_with_total = pd.concat([summary, pd.DataFrame([grand_totals])], ignore_index=True, sort=False)

    # 6) Safely write back to the same Excel file:
    #    - write to a temp copy, then replace original (reduces chance of corrupting file)
    log.info("Writing Summary sheet to workbook (safe replace)...")
    # Load existing workbook and remove existing summary sheet if present
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir) / f"{excel_path.name}.tmp.xlsx"

        # Use load_workbook to copy existing workbook then remove summary sheet if exists
        wb = load_workbook(excel_path)
        if summary_sheet in wb.sheetnames:
            ws = wb[summary_sheet]
            wb.remove(ws)
        wb.save(tmp_path)  # save a base workbook on which pandas can write

        # Now use pandas ExcelWriter to append the new summary sheet
        with pd.ExcelWriter(tmp_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            # ensure we re-open the workbook for writer.book
            writer.book = load_workbook(tmp_path)
            summary_with_total.to_excel(writer, sheet_name=summary_sheet, index=False)
            writer.save()

        # Move temp file to original path (atomic-ish replace)
        backup_path = excel_path.with_suffix(excel_path.suffix + ".bak")
        try:
            # keep a backup first
            shutil.copy2(excel_path, backup_path)
            shutil.move(str(tmp_path), str(excel_path))
            log.info("Successfully updated workbook; backup kept at: %s", backup_path)
        except Exception as e:
            log.exception("Failed to replace original workbook, restoring backup...")
            # try to restore from backup if something went wrong
            if backup_path.exists():
                shutil.copy2(backup_path, excel_path)
            raise

    log.info("Summary sheet creation complete. Summary rows: %d (incl. grand total)", len(summary_with_total))
    # Optionally return path for XCom or operators
    return str(excel_path)

# ---- DAG definition ----
with DAG(
    dag_id="create_summary_sheet",
    default_args=default_args,
    start_date=datetime(2025, 1, 1),
    schedule_interval=None,  # set to desired schedule
    catchup=False,
    tags=["excel", "summary", "pivot"],
) as dag:

    task_create_summary = PythonOperator(
        task_id="create_summary_sheet_task",
        python_callable=create_summary_sheet,
        op_kwargs={"excel_path": EXCEL_PATH_STR, "sheet_name": SHEET_NAME, "summary_sheet": SUMMARY_SHEET},
    )

    task_create_summary
